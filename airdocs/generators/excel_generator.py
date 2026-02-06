# AirDocs - Excel Generator
# =================================

import logging
import re
from pathlib import Path
from typing import Any

from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.worksheet import Worksheet

from core.exceptions import GenerationError, TemplateError
from .base_generator import BaseGenerator

logger = logging.getLogger("airdocs.generators")


class ExcelGenerator(BaseGenerator):
    """
    Generator for Excel documents using openpyxl.

    Supports two modes:
    1. Template-based: Fill placeholders in existing Excel template
    2. Programmatic: Build Excel from scratch

    Placeholder format: {{ field_name }}
    """

    TEMPLATE_TYPE = "excel"
    PLACEHOLDER_PATTERN = re.compile(r"\{\{\s*(\w+)\s*\}\}")

    def generate(
        self,
        template_name: str,
        data: dict[str, Any],
        output_path: Path | str,
    ) -> None:
        """
        Generate Excel document from template.

        Args:
            template_name: Name of the template (e.g., 'registry_1c')
            data: Data dictionary with canonical field names
            output_path: Path where to save the generated document

        Raises:
            GenerationError: If generation fails
            TemplateError: If template not found or invalid
        """
        output_path = Path(output_path)

        try:
            # Get template path
            template_path = self.get_template_path(self.TEMPLATE_TYPE, template_name)

            # Prepare context
            context = self.prepare_context(data)

            # Load template
            wb = load_workbook(template_path)

            # Process all sheets
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                self._fill_placeholders(sheet, context)

            # Ensure output directory exists
            self.ensure_output_dir(output_path)

            # Save
            wb.save(output_path)
            wb.close()

            logger.info(f"Generated Excel document: {output_path}")

        except TemplateError:
            raise
        except Exception as e:
            self._handle_generation_error(e, template_name, output_path)

    def _fill_placeholders(
        self,
        sheet: Worksheet,
        context: dict[str, Any],
    ) -> None:
        """
        Fill placeholders in a worksheet.

        Args:
            sheet: Worksheet to process
            context: Data context
        """
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    # Find and replace all placeholders
                    new_value = self._replace_placeholders(cell.value, context)
                    if new_value != cell.value:
                        # Try to preserve type if placeholder was the entire cell value
                        if self.PLACEHOLDER_PATTERN.fullmatch(cell.value.strip()):
                            # Single placeholder - try to use original type
                            cell.value = new_value
                        else:
                            cell.value = new_value

    def _replace_placeholders(
        self,
        text: str,
        context: dict[str, Any],
    ) -> Any:
        """
        Replace placeholders in text with values from context.

        Args:
            text: Text containing placeholders
            context: Data context

        Returns:
            Text with placeholders replaced, or the value if single placeholder
        """
        # Check if it's a single placeholder (entire cell value)
        match = self.PLACEHOLDER_PATTERN.fullmatch(text.strip())
        if match:
            field_name = match.group(1)
            return context.get(field_name, "")

        # Multiple placeholders or text with placeholders
        def replace(m):
            field_name = m.group(1)
            value = context.get(field_name, "")
            return str(value) if value is not None else ""

        return self.PLACEHOLDER_PATTERN.sub(replace, text)

    def generate_registry(
        self,
        data: list[dict[str, Any]],
        output_path: Path | str,
        columns: list[tuple[str, str]] | None = None,
    ) -> None:
        """
        Generate a registry Excel file from a list of records.

        Args:
            data: List of data dictionaries
            output_path: Path where to save
            columns: Optional list of (field_name, header) tuples
        """
        output_path = Path(output_path)

        # Default columns for 1C registry
        if columns is None:
            columns = [
                ("awb_number", "AWB №"),
                ("shipment_date", "Дата"),
                ("shipper_name", "Отправитель"),
                ("consignee_name", "Получатель"),
                ("weight_kg", "Вес (кг)"),
                ("pieces", "Мест"),
                ("goods_description", "Описание"),
            ]

        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Реестр"

            # Write headers
            for col_idx, (field, header) in enumerate(columns, start=1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.font = cell.font.copy(bold=True)

            # Write data rows
            for row_idx, record in enumerate(data, start=2):
                for col_idx, (field, _) in enumerate(columns, start=1):
                    value = record.get(field, "")
                    ws.cell(row=row_idx, column=col_idx, value=value)

            # Auto-adjust column widths (basic)
            for col_idx, (field, header) in enumerate(columns, start=1):
                ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = max(12, len(header) + 2)

            # Ensure output directory exists
            self.ensure_output_dir(output_path)

            wb.save(output_path)
            wb.close()

            logger.info(f"Generated registry Excel: {output_path}")

        except Exception as e:
            raise GenerationError(
                f"Failed to generate registry: {e}",
                document_type="registry",
                cause=e,
            )

    def generate_from_file(
        self,
        template_path: Path | str,
        data: dict[str, Any],
        output_path: Path | str,
    ) -> None:
        """
        Generate Excel from a specific template file path.

        Args:
            template_path: Full path to the template file
            data: Data dictionary
            output_path: Path where to save
        """
        template_path = Path(template_path)
        output_path = Path(output_path)

        if not template_path.exists():
            raise TemplateError(
                f"Template file not found: {template_path}",
                template_path=str(template_path),
            )

        try:
            context = self.prepare_context(data)
            wb = load_workbook(template_path)

            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                self._fill_placeholders(sheet, context)

            self.ensure_output_dir(output_path)
            wb.save(output_path)
            wb.close()

            logger.info(f"Generated Excel from custom template: {output_path}")

        except Exception as e:
            self._handle_generation_error(e, str(template_path), output_path)

    def get_template_fields(self, template_name: str) -> list[str]:
        """
        Get list of placeholder fields in a template.

        Args:
            template_name: Name of the template

        Returns:
            List of field names found in template
        """
        fields = set()

        try:
            template_path = self.get_template_path(self.TEMPLATE_TYPE, template_name)
            wb = load_workbook(template_path)

            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                for row in sheet.iter_rows():
                    for cell in row:
                        if cell.value and isinstance(cell.value, str):
                            matches = self.PLACEHOLDER_PATTERN.findall(cell.value)
                            fields.update(matches)

            wb.close()
            return list(fields)

        except Exception as e:
            logger.warning(f"Could not extract fields from template: {e}")
            return []
